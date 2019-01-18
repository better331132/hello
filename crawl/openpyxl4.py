import openpyxl
from openpyxl.styles.borders import Border, Side
from PIL import Image
book = openpyxl.Workbook()
sheet = book.active
tmpCell = sheet['C1']
# tmpCell = sheet1.cell(row=1, column=3)
tmpCell.font = openpyxl.styles.Font(size=14, color='FF0000')
tmpCell.number_format
tmpCell.value = 12345
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

cell = sheet.cell(row=3, column=2)
cell.border = thin_border
# insert image
imgFile = '../crawl/images/test.png'
img = openpyxl.drawing.image.Image(imgFile)
sheet.add_image(img, 'B5')

book.save("./data/openpyxl4.xlsx")

img2 = Image.open(imgFile)
new_img = img2.resize((100, 100))
new_img.save('new.png')
img3 = openpyxl.drawing.image.Image('new.png')
sheet.add_image(img3, 'A5')