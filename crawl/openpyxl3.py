import openpyxl
from openpyxl.styles.borders import Border, Side
book = openpyxl.Workbook()
sheet = book.active
tmpCell = sheet['C1']
# tmpCell = sheet1.cell(row=1, column=3)
tmpCell.font = openpyxl.styles.Font(size=14, color='FF0000')
tmpCell.number_format
tmpCell.value = 12345
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

cell = sheet.cell(row=3, column=2)
cell.border = thin_border
book.save("./data/openpyxl3.xlsx")


#Value must be one of {‘double’, ‘mediumDashed’, ‘mediumDashDotDot’, ‘hair’, ‘dashDot’, ‘dashDotDot’, ‘dashed’, ‘mediumDashDot’, ‘medium’, ‘slantDashDot’, ‘thin’, ‘thick’, ‘dotted’}