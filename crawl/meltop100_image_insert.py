import requests
from bs4 import BeautifulSoup
import urllib.request
import openpyxl
import csv

wb = openpyxl.Workbook()
ws2 = wb.create_sheet()


for i in range(1,101):
    imgFile = './images/meltop100/meltop{}.png'.format(i)
    img = openpyxl.drawing.image.Image(imgFile)
    tmpCell = ws2.cell(row = 6 * i - 5, column = 1)
    tmpCell.value = "Top{} Image".format(i)
    ws2.add_image(img, 'C{}'.format(6 * i - 5))
    print("OK === >> {}".format(i))


wb.save('./data/melon_top100_excel.xlsx')

