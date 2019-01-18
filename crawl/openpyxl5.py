import openpyxl
from openpyxl.styles.borders import Border, Side
from PIL import Image
from openpyxl.chart import (
    Reference,
    BarChart
)
from openpyxl.chart import (
    Reference, Series,
    ScatterChart
)

book = openpyxl.Workbook()
sheet = book.active
tmpCell = sheet['C1']
# tmpCell = sheet1.cell(row=1, column=3)
tmpCell.font = openpyxl.styles.Font(size=14, color='FF0000')
tmpCell.number_format
tmpCell.value = 12345
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

cell = sheet.cell(row=3, column=2)
cell.border = thin_border
# insert image
imgFile = '../crawl/images/test.png'
img = openpyxl.drawing.image.Image(imgFile)
sheet.add_image(img, 'B5')

book.save("./data/openpyxl4.xlsx")

img2 = Image.open(imgFile)
new_img = img2.resize((100, 100))
new_img.save('new.png')
img3 = openpyxl.drawing.image.Image('new.png')
sheet.add_image(img3, 'A5')

book.save("./data/openpyxl5.xlsx")

# rows = [
#     ['김일수', 11],
#     ['김이수', 22],
#     ['김삼수', 33],
#     ['김사수', 15],
#     ['김오수', 11],
# ]

# for row in rows:
#     sheet.append(row)

# datax = Reference(sheet, min_col=2, 
# 		min_row=1, max_col=2, max_row=5)
# categs = Reference(sheet, min_col=1,
# 				 min_row=1, max_row=5)

# chart = BarChart()
# chart.add_data(data=datax)
# chart.set_categories(categs)

# chart.legend = None  # 범례
# chart.varyColors = True
# chart.title = "차트 타이틀"

# sheet.add_chart(chart, "A8")

# rows = [
#     ['Size', 'Batch 1', 'Batch 2'],
#     [2, 40, 30],
#     [3, 40, 25],
#     [4, 50, 30],
#     [5, 30, 25],
#     [6, 25, 35],
#     [7, 20, 40],
# ]

# for row in rows:
#     sheet.append(row)

# chart = ScatterChart()
# chart.style = 13
# chart.x_axis.title = 'Size'
# chart.y_axis.title = 'Percentage'

# xvalues = Reference(ws, min_col=1,
# 			 min_row=2, max_row=7)

# for i in range(2, 4):
#     values = Reference(ws, 
# 				min_col=i, 
# 				min_row=1, 
# 				max_row=7)
#     series = Series(values, xvalues, 
# 				title_from_data=True)
#     chart.series.append(series)

# ws.add_chart(chart, "A10")