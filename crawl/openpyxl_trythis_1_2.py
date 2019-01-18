import requests
from bs4 import BeautifulSoup
import urllib.request
import openpyxl
import csv


wb = openpyxl.Workbook()
ws = wb.active
ws.title = "챠트목록"
ws2 = wb.create_sheet()
ws2.title = "Image 목록"

with open('./data/meltop100.csv', 'r', encoding='utf-8') as f:
    for num, row in enumerate(csv.reader(f)):
        if num == 0:
            ws.append(row)   
        elif num <= 100:
            ws.append([int(row[0]),row[1],row[2],int(row[3]), int(row[4])])
        else:
            ws.append([row[0], row[1], row[2], int(row[3]), int(row[4])])

for i in range(1,101):
    imgFile = './images/meltop100/meltop{}.png'.format(i)
    img = openpyxl.drawing.image.Image(imgFile)
    tmpCell = ws2.cell(row = 6 * i - 5, column = 1)
    tmpCell.value = "Top{} Image".format(i)
    ws2.add_image(img, 'C{}'.format(6 * i - 5))
    print("OK === >> {}".format(i))


wb.save('./data/melon_top100_excel.xlsx')