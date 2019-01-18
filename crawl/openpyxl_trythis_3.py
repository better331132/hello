import requests
from bs4 import BeautifulSoup
import urllib.request
import openpyxl
import csv
from openpyxl.chart import (
    Reference,
    BarChart
)
from openpyxl.chart import (
    Reference, Series,
    ScatterChart
)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "챠트목록"
ws2 = wb.create_sheet()
ws2.title = "Image 목록"
ws3 = wb.create_sheet()
ws3.title = "Top10 Chart"

with open('./data/meltop100.csv', 'r', encoding='utf-8') as f:
    for num, row in enumerate(csv.reader(f)):
        if num == 0:
            ws.append(row)   
        elif num <= 100:
            ws.append([int(row[0]),row[1],row[2],int(row[3]), int(row[4])])
        else:
            ws.append([row[0], row[1], row[2], int(row[3]), int(row[4])])

for i in range(1,101):
    imgFile = './images/meltop100/meltop{}.png'.format(i)
    img = openpyxl.drawing.image.Image(imgFile)
    tmpCell = ws2.cell(row = 6 * i - 5, column = 1)
    tmpCell.value = "Top{} Image".format(i)
    ws2.add_image(img, 'C{}'.format(6 * i - 5))
    print("OK === >> {}".format(i))


dat = Reference(ws, min_col=4, min_row=2, max_col=4, max_row=11)
cat = Reference(ws, min_col=2, min_row=2, max_col=2, max_row=11)


barChart = BarChart()
barChart.add_data(data=dat)
barChart.set_categories(cat)

barChart.legend = None  # 범례
barChart.varyColors = True
barChart.title = "Top10 Likes"

ws3.add_chart(barChart, "A1")

scatChart = ScatterChart()
scatChart.style = 13
scatChart.x_axis.title = 'SongName'
scatChart.y_axis.title = 'DiffLikes'

xvalues = Reference(ws, min_col=1,
			 min_row=2, max_row=11)

lst = [5]
for i in lst:
    values = Reference(ws, 
				min_col=i,
                min_row=1,
                max_row=11)
    series = Series(values, xvalues, 
				title_from_data=True)
    scatChart.series.append(series)
    
ws3.add_chart(scatChart, "A20")

print("됨")
wb.save('./data/melon_top100_excel.xlsx')